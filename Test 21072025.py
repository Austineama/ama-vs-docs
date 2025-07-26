# Import necessary libraries
import os
from posixpath import basename
import pandas as pd
import logging
import glob
import openpyxl


# --- 1. Configure Logging ---# Change to DEBUG for more detail, WARNING/ERROR for less
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s",)

# --- 2. Define File Paths ---
folder = r"D:\Ama-VSProject"
file_pattern = "*.xlsx"
excel_files = glob.glob(os.path.join(folder, file_pattern))


for file in excel_files:
    try:
        wb = openpyxl.load_workbook(file)
        print(f"Sheets in {file}:")
        for sheet in wb.sheetnames:
            print(sheet)
        print()
    except Exception as e:
        print(f"Error opening file {file}: {e}")











filepath = os.path.join(folder, file_pattern)

excel_files = glob.glob(filepath)

for file in excel_files:
    wb = openpyxl.load_workbook(file)
    print(f"Sheets in {file}:")
    for sheet in wb.sheetnames:
        print(sheet)
    print()

File_name = [basename(file) for file in excel_files]
print(File_name)

# Log the file path
logging.info(f"File path set to: {filepath}")

print("Filename:", file_1)
print("Full filepath:", filepath)

# Check if the file exists
if not os.path.exists(filepath):
    logging.error(f"File not found: {filepath}")
    raise FileNotFoundError(f"File not found: {filepath}")
# If the file exists, proceed with loading
logging.info(f"File found: {filepath}")

# Replace 'your_file.xlsx' with the path to your Excel file
excel_file = file_1 

# Create an ExcelFile object
Df3 = pd.ExcelFile(filepath)

# Get the list of sheet names
sheet_names = Df3.sheet_names

print(sheet_names)


# --- 2. Data Loading Functions ---
/*************  ✨ Windsurf Command 🌟  *************/
def load_excel_sheet(filepath: str, sheet_name: str) -> pd.DataFrame:
    """
    Load a specific sheet from an Excel file.

    Args:
        file_path (str): Path to the Excel file.
        filepath (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to load.

    Returns:
        pd.DataFrame: The loaded data.

    Raises:
        FileNotFoundError: If the file does not exist.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    return pd.read_excel(file_path, sheet_name=sheet_name)
/*******  80a1643f-7eea-4e7d-81a2-ebe0442f3cde  *******/        

# Load a specific sheet from the Excel file
df = pd.read_excel(filepath, sheet_name=sheet_name)

Activity_df = pd.read_excel(filepath, sheet_name=1)  # Load specific sheet

# Load all sheets from the Excel file
BP_Activity_df = pd.read_excel(filepath, sheet_name=None)  # Load all sheets

excel_file = pd.ExcelFile(filepath)
sheet_names = excel_file.sheet_names

print(sheet_names)
print("Excel worksheet names:", sheet_names)

if not os.path.exists(filepath):
    logging.error(f"File not found: {filepath}")
    raise FileNotFoundError(f"File not found: {filepath}")

if sheet_name is None:
    logging.info(f"Loading all sheets from '{filepath}'...")
else:
    logging.info(f"Loading sheet '{sheet_name}' from '{filepath}'...")


# Define parameter and function to load all sheets from an Excel file
df.to_excel(columns_file_path, index=False, sheet_name="Columns")

df = pd.read_excel(file_path, sheet_name=sheet_name)


# --- 3. Processing Functions ---
def filter_high_steps(df: pd.DataFrame, min_steps: int = 12000) -> pd.DataFrame:
    """Return only rows where Total steps > min_steps."""
    logging.info(f"Filtering rows: Total steps > {min_steps}")
    return df[df["Total steps"] > min_steps]


def merge_on_date(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """Merge two DataFrames on the 'Date' column."""
    logging.info(f"Merging DataFrames on 'Date'")
    return pd.merge(df1, df2, on="Date")


# --- 4. Output Functions ---
def save_to_csv(df: pd.DataFrame, filename: str):
    """Save DataFrame to CSV."""
    logging.info(f"Saving DataFrame to '{filename}'")
    df.to_csv(filename, index=False)


# --- 5. Main Pipeline Function ---
def run_pipeline(activity_path: str, bp_path: str, output_path: str):
    # Step 1: Load data
    activity_df = load_excel_sheet(activity_path, sheet_name="Activity")
    bp_df = load_excel_sheet(bp_path, sheet_name="BP Readings")

    # Step 2: Filter for high step count
    high_steps_df = filter_high_steps(activity_df)

    # Step 3: Merge activity with BP readings
    merged_df = merge_on_date(high_steps_df, bp_df)

    # Step 4: Save results
    save_to_csv(merged_df, output_path)
    logging.info("Pipeline completed successfully.")


# --- 6. Script Entry Point ---
if __name__ == "__main__":
    ACTIVITY_PATH = "BP_Activity.xlsx"  # Path to your Excel file (Activity)
    BP_PATH = "BP_Activity.xlsx"  # Path to your Excel file (BP Readings)
    OUTPUT_PATH = "filtered_activity_bp.csv"
    run_pipeline(ACTIVITY_PATH, BP_PATH, OUTPUT_PATH)


folder = r"D:\Ama-VSProject"
file_pattern = "BP_Activity.xlsx|Plan_1_Results.xlsx"  # pattern to match specific files
filepath = os.path.join(folder, file_pattern)

excel_files = glob.glob(filepath)

excel_files = glob.glob(filepath)
print(excel_files)